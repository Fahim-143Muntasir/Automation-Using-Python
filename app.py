import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb=xl.load_workbook("transactions.xlsx")
sheet=wb["Sheet1"]
for row in range(2, sheet.max_row+1):
    cell= sheet.cell(row,3)
    updated_price=cell.value*.9
    new_cell= sheet.cell(row,4)
    new_cell.value=updated_price
wb.save('transactions2.xlsx')

